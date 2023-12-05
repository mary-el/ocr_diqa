import multiprocessing as mp
import pickle
from pathlib import Path

import cv2
import matplotlib.pyplot as plt
import numpy as np
import tqdm
import numpy.lib.recfunctions as rfn
import pandas as pd
import openpyxl

from src.data.parsed_image import ParsedImage
from src.settings import N_PROC, RAW_DATA_PATH
from src.utils.common import resize


def process_file(file, width, pred):
    img = cv2.imread(str(file.with_suffix('.jpg')))
    img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    img = resize(img, width)
    parsed = ParsedImage(img)
    gt = np.array((pred), dtype=[('gt', float)])
    series = rfn.merge_arrays((parsed.series, gt), flatten=True)
    return series


def create_requests_ds(path: Path, save_file: str, size: int = 1024) -> None:
    images_folder_paths = [path / 'court', path / 'mvd']
    gt_files = [path / 'court_gt.xlsx', path / 'mvd_gt.xlsx']
    df = []
    gt = []
    for i, image_folder_path in enumerate(images_folder_paths):
        gt_file = gt_files[i]
        wb = openpyxl.load_workbook(gt_file)
        ws = wb.worksheets[1]
        for j, cell in enumerate(ws['A'][1:], 1):
            gt.append((image_folder_path / cell.value, float(ws['D'][j].value) / 100))
    with mp.Pool(N_PROC) as pool:
        df.extend(pool.starmap(process_file, tqdm.tqdm([(file, size, pred) for file, pred in gt])))
    df = np.stack(df)
    with open(save_file, 'wb') as f:
        pickle.dump(df, f)


if __name__ == '__main__':
    create_requests_ds(Path('data'), 'data/req.pkl')