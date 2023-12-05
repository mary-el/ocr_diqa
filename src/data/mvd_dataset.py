import os
import pickle
from pathlib import Path

import cv2
import matplotlib.pyplot as plt
import numpy as np
import numpy.lib.recfunctions as rfn
import openpyxl
import tqdm
from src.data.parsed_image import ParsedImage
from src.utils.common import resize


def create_mvd_dataset(data_path: Path, save_file, width=1024):
    wb = openpyxl.load_workbook('data/mvd.xlsx')
    ws = wb.active
    arr = []
    for i, cell in enumerate(ws['C'][1:], 1):
        if ws['D'][i].value == 0:
            arr.append((cell.value, float(ws['E'][i].value == ws['G'][i].value), float(ws['J'][i].value == 'хк')))
    df = []
    for file, pred, sub in tqdm.tqdm(arr):
        series = process_file(data_path, file, width)
        gt = np.array((pred, sub), dtype=[('pred', float), ('sub', float)])
        series = rfn.merge_arrays((series, gt), flatten=True)
        df.append(series)
    df = np.stack(df)
    with open(save_file, 'wb') as f:
        pickle.dump(df, f)


def process_file(data_path, file, width):
    if file.startswith('cl_mvd_zakaz_'):
        file = file[len('cl_mvd_zakaz_'):]
    img = cv2.imread(str((data_path / file).with_suffix('.jpg')))
    img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    img = resize(img, width)
    parsed = ParsedImage(img)
    return parsed.series


if __name__ == '__main__':
    create_mvd_dataset(Path('data/mvd'), 'data/mvd.pkl')
